import requests
import datetime
from datetime import timedelta
import telebot
from openpyxl import load_workbook


def open_exel():
    book = load_workbook("Table_SKU.xlsx", read_only=True)
    sheet = book.active
    max_sheet = sheet.max_row
    result = []
    for i in range(1, max_sheet + 1):
        result.append(sheet[i][0].value)
    return result


def imtid_sku(SKU):
    src = requests.get(f'https://wbx-content-v2.wbstatic.net/ru/{SKU}.json')
    data = src.json()
    imt_id = data['imt_id']
    return imt_id


def text_comment(rezult, SKU, message):
    info_comments = []
    for item in range(len(rezult)):
        date = rezult[item]['createdDate']
        date_time_obj = datetime.datetime.strptime(date, "%Y-%m-%dT%H:%M:%SZ")
        date_time = date_time_obj.date()
        date_com = datetime.date.today() - timedelta(days=1)
        if date_time >= date_com:
            if rezult[item]['productValuation'] < 5:
                text = 'Негативный отзыв'
                info_comments.append(text)
                name = rezult[item]['wbUserDetails']['name']
                info_comments.append(name)
                name_detail = rezult[item]['productDetails']['productName']
                info_comments.append(name_detail)
                info_comments.append(f'Номер карточки {SKU}')
                stars = rezult[item]['productValuation']
                info_comments.append(f'Оценка пользователя {stars}')
                text_comments = rezult[item]['text']
                info_comments.append(f"\n{text_comments}")
                date = rezult[item]['createdDate']
                date_time_obj = datetime.datetime.strptime(date, "%Y-%m-%dT%H:%M:%SZ")
                info_comments.append(f'\n{date_time_obj}')
                comments = '|'.join(str(x) for x in info_comments)
                bot.send_message(chat_id=message.chat.id, text=comments)
                info_comments.clear()
        else:
            break


def post_data(imt_id, i):
    headers = {
        'accept': '*/*',
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.61 Safari/537.36',
    }
    json_data = {
        'imtId': imt_id,
        'skip': i,
        'take': 20,
        'order': 'dateDesc',
    }
    response = requests.post('https://public-feedbacks.wildberries.ru/api/v1/feedbacks/site', headers=headers,
                             json=json_data)
    src = response.json()
    return src





@bot.message_handler(commands=["start"])
def handler(message):
    result = open_exel()
    for SKU in result:
        imt_id = imtid_sku(SKU)
        i = 0
        while True:
            src = post_data(imt_id, i)
            i += 20
            rezult = src['feedbacks']
            text_comment(rezult, SKU, message)
            if i == 300:
                break


if __name__ == '__main__':
    bot.polling(non_stop=True)
